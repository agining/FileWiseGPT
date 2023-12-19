import textract
import openpyxl
from PyPDF2 import PdfReader

class FileProcessor:
    def excel_to_text(file_path):
        # Convert Excel file to text
        wb = openpyxl.load_workbook(file_path)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        text += str(cell) + " "
        return text

    def word_to_text(file_path):
        # Convert Word document to text
        text = textract.process(file_path).decode("utf-8")
        return text

    def pdf_to_text(file_path):
        # Convert PDF file to text
        text = ""
        pdf_reader = PdfReader(file_path)
        for page in pdf_reader.pages:
            text += page.extract_text()
        return text